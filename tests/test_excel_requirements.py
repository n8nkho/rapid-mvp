"""
Tests for Excel upload/download of requirements (Sprint 6).
"""
from io import BytesIO
import os
import sys
from unittest.mock import patch

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from main import app  # noqa: E402


ENGAGEMENT = "eng-excel-001"


def _make_req(req_id, **overrides):
    base = {
        "req_id": req_id,
        "engagement_id": ENGAGEMENT,
        "title": f"Requirement {req_id}",
        "description": "Desc",
        "business_process": "Record-to-Report",
        "priority": "Must-Have",
        "category": "Automation",
        "tags": ["pain_point"],
        "stakeholder": "CFO",
        "sign_off_status": "draft",
    }
    base.update(overrides)
    return base


def _client():
    return TestClient(app, raise_server_exceptions=True)


class TestExportExcel:
    def test_export_basic_structure(self):
        client = _client()
        reqs = [
            _make_req("REQ-001", tags=["pain_point", "manual_step"]),
            _make_req("REQ-002", tags=[]),
        ]
        with patch("main.get_requirements_by_engagement", return_value=reqs):
            resp = client.get(f"/engagement/{ENGAGEMENT}/requirements/export")

        assert resp.status_code == 200
        assert resp.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))[0:10]]
        assert "req_id" in headers
        assert "title" in headers
        assert "tags" in headers

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert rows[0][0] == "REQ-001"
        assert "pain_point" in (rows[0][7] or "")


class TestImportExcel:
    def _build_workbook(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Requirements"
        ws.append(
            [
                "req_id",
                "title",
                "description",
                "business_process",
                "priority",
                "category",
                "tags",
                "stakeholder",
                "sign_off_status",
            ]
        )
        ws.append(
            [
                "",  # new requirement
                "From Excel",
                "Created from Excel sheet",
                "Order-to-Cash",
                "Should-Have",
                "Integration",
                "Pain point, Manual step",
                "CFO",
                "draft",
            ]
        )
        return wb

    def test_import_creates_requirement_and_normalises_tags(self):
        client = _client()
        wb = self._build_workbook()
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        created = _make_req("REQ-010", tags=["pain_point", "manual_step"])

        with (
            patch("main.get_requirements_by_engagement", return_value=[]),
            patch("main.create_requirement", return_value=created) as mock_create,
        ):
            resp = client.post(
                f"/engagement/{ENGAGEMENT}/requirements/import",
                files={
                    "file": (
                        "reqs.xlsx",
                        buf.getvalue(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )

        assert resp.status_code == 200
        data = resp.json()
        assert data["created"] == 1
        assert "REQ-010" in data["created_req_ids"]

        call_kwargs = mock_create.call_args.kwargs
        assert call_kwargs["engagement_id"] == ENGAGEMENT
        assert set(call_kwargs["tags"]) == {"pain_point", "manual_step"}
        assert call_kwargs["business_process"] == "Order-to-Cash"
        assert call_kwargs["priority"] == "Should-Have"
        assert call_kwargs["category"] == "Integration"

    def test_import_updates_existing_requirement_when_req_id_present(self):
        client = _client()
        wb = Workbook()
        ws = wb.active
        ws.title = "Requirements"
        ws.append(["req_id", "title", "description", "tags"])
        ws.append(["REQ-001", "Updated title", "Updated desc", "Pain point"])

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        existing = [_make_req("REQ-001")]
        updated = _make_req("REQ-001", title="Updated title", description="Updated desc", tags=["pain_point"])

        with (
            patch("main.get_requirements_by_engagement", return_value=existing),
            patch("main.update_requirement", return_value=updated) as mock_update,
        ):
            resp = client.post(
                f"/engagement/{ENGAGEMENT}/requirements/import",
                files={
                    "file": (
                        "reqs.xlsx",
                        buf.getvalue(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )

        assert resp.status_code == 200
        data = resp.json()
        assert data["updated"] == 1
        assert "REQ-001" in data["updated_req_ids"]
        # Ensure tags were passed as normalised list
        call_kwargs = mock_update.call_args.kwargs
        assert call_kwargs["req_id"] == "REQ-001"
        assert call_kwargs["engagement_id"] == ENGAGEMENT
        assert call_kwargs["updates"]["title"] == "Updated title"
        assert call_kwargs["updates"]["description"] == "Updated desc"
        assert call_kwargs["updates"]["tags"] == ["pain_point"]

    def test_import_rejects_non_excel(self):
        client = _client()
        resp = client.post(
            f"/engagement/{ENGAGEMENT}/requirements/import",
            files={"file": ("reqs.csv", b"not excel", "text/csv")},
        )
        assert resp.status_code == 400

